# -*- coding: utf-8 -*-
"""
Complementa la información de productos del catálogo con peso, dimensiones (largo, ancho, alto)
y/o stock desde un archivo Excel.

Formatos soportados:
  - Hoja con cabecera: columna A = código/identificador, luego columnas peso, largo, ancho, alto
    (nombres de cabecera en español: peso, largo, ancho, alto; opcional: ESTOCK o stock).
  - Hoja tipo CAT: código (ej. ET001), ESTOCK, y columnas L/largo, ancho, alto.

Solo actualiza productos que ya existan en el catálogo (por SKU o por coincidencia de nombre).
No crea productos nuevos.

Uso:
  python manage.py load_product_specs "ruta/al/archivo.xlsx"
  python manage.py load_product_specs "ruta/al/archivo.xlsx" --sheet "Hoja1"
  python manage.py load_product_specs "ruta/al/archivo.xlsx" --dry-run

Requisito: pip install openpyxl
"""
import re
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from apps.catalog.models import Product


def _to_decimal(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(",", ".")
    try:
        n = re.sub(r"[^\d.-]", "", s) or "0"
        d = Decimal(n)
        return d if d != 0 else None
    except Exception:
        return None


def _to_int(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    try:
        return int(float(str(val).strip().replace(",", ".")))
    except (ValueError, TypeError):
        return None


def _normalize_identifier(val):
    """Limpia el identificador para búsqueda (SKU o nombre)."""
    if val is None:
        return ""
    s = str(val).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _find_product(identifier):
    """
    Busca un producto por identificador: SKU exacto, SKU normalizado (sin espacios/guiones),
    o por nombre que contenga el texto.
    """
    if not identifier:
        return None
    id_clean = _normalize_identifier(identifier)
    id_upper = id_clean.upper().strip()
    id_sku_like = re.sub(r"[\s\-]+", "", id_upper)[:50]

    # 1. SKU exacto (case insensitive)
    p = Product.objects.filter(sku__iexact=id_upper, is_active=True).first()
    if p:
        return p
    # 2. SKU sin espacios ni guiones
    if id_sku_like:
        p = Product.objects.filter(sku__iexact=id_sku_like, is_active=True).first()
        if p:
            return p
    # 3. Nombre contiene el identificador (útil para "twg 10.7 euro 3")
    if id_clean:
        p = Product.objects.filter(name__icontains=id_clean, is_active=True).first()
        if p:
            return p
    return None


def _detect_headers(rows):
    """
    Encuentra la fila de cabecera y un mapa col_index -> nombre normalizado.
    Retorna (header_row_index, {col_idx: 'peso'|'largo'|'ancho'|'alto'|'stock'|'codigo'}, first_data_row).
    """
    # Nombres aceptados por campo
    PESO = {"peso", "weight"}
    LARGO = {"largo", "length", "l ", "largo "}
    ANCHO = {"ancho", "width", "aπ", "a "}
    ALTO = {"alto", "height", "al ", "altura"}
    STOCK = {"estock", "stock"}
    CODIGO = {"codigo", "código", "sku", "code", "part#", "parte"}

    def norm(h):
        if h is None:
            return ""
        return re.sub(r"\s+", " ", str(h).strip().lower())[:30]

    for row_idx, row in enumerate(rows):
        if not row:
            continue
        row = list(row) if row else []
        mapping = {}
        has_spec = False
        for col_idx, cell in enumerate(row):
            h = norm(cell)
            if not h:
                continue
            if h in PESO or "peso" in h:
                mapping[col_idx] = "peso"
                has_spec = True
            elif h in ALTO or h.startswith("al ") or "alto" in h:
                mapping[col_idx] = "alto"
                has_spec = True
            elif h in LARGO or h.startswith("l ") or "largo" in h:
                mapping[col_idx] = "largo"
                has_spec = True
            elif h in ANCHO or "aπ" in h or "ancho" in h or (h.startswith("a") and "ncho" in h):
                mapping[col_idx] = "ancho"
                has_spec = True
            elif h in STOCK or "stock" in h:
                mapping[col_idx] = "stock"
                has_spec = True
            elif h in CODIGO or col_idx == 0:
                mapping[col_idx] = "codigo"
        if has_spec:
            # La primera columna con dato suele ser el código/identificador
            if "codigo" not in mapping.values():
                mapping[0] = "codigo"
            return row_idx, mapping, row_idx + 1
    return -1, {}, 0


def _parse_sheet(ws):
    """
    Parsea una hoja y devuelve lista de dicts con keys: codigo, peso, largo, ancho, alto, stock.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row, col_map, start = _detect_headers(rows)
    if header_row < 0 or not col_map:
        return []

    data = []
    for row in rows[start:]:
        row = list(row) if row else []
        if not row:
            continue
        codigo = None
        for idx, name in col_map.items():
            if name == "codigo" and idx < len(row):
                codigo = row[idx]
                break
        if codigo is None and len(row) > 0:
            codigo = row[0]
        codigo = _normalize_identifier(codigo)
        if not codigo:
            continue

        item = {"codigo": codigo, "peso": None, "largo": None, "ancho": None, "alto": None, "stock": None}
        for idx, name in col_map.items():
            if name == "codigo" or idx >= len(row):
                continue
            val = row[idx]
            if name == "peso":
                item["peso"] = _to_decimal(val)
            elif name == "largo":
                item["largo"] = _to_decimal(val)
            elif name == "ancho":
                item["ancho"] = _to_decimal(val)
            elif name == "alto":
                item["alto"] = _to_decimal(val)
            elif name == "stock":
                item["stock"] = _to_int(val)

        data.append(item)
    return data


class Command(BaseCommand):
    help = "Complementa productos del catálogo con peso, dimensiones y stock desde un Excel."

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            type=str,
            help="Ruta al archivo .xlsx",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default=None,
            help="Nombre de la hoja a leer (por defecto todas las hojas).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Solo mostrar qué se actualizaría, sin guardar.",
        )

    def handle(self, path, sheet, dry_run, **options):
        path = Path(path).resolve()
        if not path.exists():
            self.stderr.write(self.style.ERROR(f"Archivo no encontrado: {path}"))
            return

        self.stdout.write(f"Leyendo {path} ...")
        wb = openpyxl.load_workbook(path, read_only=True)

        sheet_names = [sheet] if sheet else wb.sheetnames
        if sheet and sheet not in wb.sheetnames:
            self.stderr.write(self.style.ERROR(f"Hoja '{sheet}' no existe. Hojas: {wb.sheetnames}"))
            wb.close()
            return

        updated = 0
        not_found = []

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            items = _parse_sheet(ws)
            if not items:
                self.stdout.write(f"  {sheet_name}: no se detectó cabecera peso/largo/ancho/alto o sin filas.")
                continue

            self.stdout.write(f"  {sheet_name}: {len(items)} filas de datos.")
            for item in items:
                product = _find_product(item["codigo"])
                if not product:
                    not_found.append(item["codigo"])
                    continue

                updates = {}
                if item["peso"] is not None:
                    updates["weight"] = item["peso"]
                if item["largo"] is not None:
                    updates["length"] = item["largo"]
                if item["ancho"] is not None:
                    updates["width"] = item["ancho"]
                if item["alto"] is not None:
                    updates["height"] = item["alto"]
                if item["stock"] is not None:
                    updates["stock"] = item["stock"]

                if not updates:
                    continue

                if dry_run:
                    self.stdout.write(
                        f"    [DRY] {product.sku}: {updates}"
                    )
                else:
                    for key, value in updates.items():
                        setattr(product, key, value)
                    product.save(update_fields=list(updates.keys()))
                updated += 1

        wb.close()

        if not_found and len(not_found) <= 25:
            self.stdout.write(self.style.WARNING(f"Sin coincidencia en catálogo ({len(not_found)}): {not_found[:15]}{'...' if len(not_found) > 15 else ''}"))
        elif not_found:
            self.stdout.write(self.style.WARNING(f"Sin coincidencia en catálogo: {len(not_found)} códigos."))

        if dry_run:
            self.stdout.write(self.style.SUCCESS(f"DRY RUN: se actualizarían {updated} productos. Ejecuta sin --dry-run para guardar."))
        else:
            self.stdout.write(self.style.SUCCESS(f"Listo: {updated} productos actualizados con peso/dimensiones/stock."))
