# -*- coding: utf-8 -*-
"""
Carga productos desde lista precios publico.xlsx al catálogo.
Una hoja = una categoría. Productos se asignan a categorías/subcategorías existentes.
Flexibles: se añade "Reforzado" al nombre.
Uso: python manage.py load_precios_xlsx "ruta/al/archivo.xlsx"

Requisito: pip install openpyxl  (o: pip install -r requirements-catalog.txt)
"""
import re
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.utils.text import slugify

from apps.catalog.models import Category, Product
from apps.catalog.flexibles_nomenclature import get_display_name_for_sku

# Mapeo: nombre normalizado de hoja Excel -> slug de categoría destino (o parent para TWC)
# Permite variantes con espacios dobles o mayúsculas/minúsculas
SHEET_TO_CATEGORY_SLUG = {
    "silenciadores": "silenciadores",
    "flexibles": "flexibles",
    "cataliticos clf": "cataliticos-twc",
    "cataliticos  clf": "cataliticos-twc",
    "cataliticos twc": "cataliticos-twc",
    "cataliticos tipo original": "cataliticos-ensamble-directo",
    "colas de escape": "colas-de-escape",
    "resonadores": "resonadores",
}

# Sufijo para todos los productos de la hoja Flexibles
FLEXIBLES_SUFFIX = " Reforzado"

# Silenciadores = marcas LTM, GW. Resonadores = LTM (por part number).
# Part numbers que se consideran RESONADORES; el resto de la hoja SILENCIADORES = Silenciadores.
# Prefijos típicos resonadores LTM (serie 08/09/10); Silenciadores LTM (ej. LTM12xx), GW, DW, DWR, LT.
# Ajustar RESONATOR_* según tu catálogo si hace falta.
RESONATOR_SKU_PREFIXES = (
    "LTM08", "LTM09", "LTM10", "LTM08-", "LTM09-", "LTM10-",
    "LTM-RES", "GW-RES", "RES-", "RESON",
)
RESONATOR_SKU_CONTAINS = ("RESON", "-RES", " RES ")


def _normalize_category_name(sheet_name):
    """Limpia nombre de hoja para usarlo como categoría."""
    name = (sheet_name or "").strip()
    name = re.sub(r"\s+", " ", name)
    return name.title() if name else "Sin categoría"


def _sheet_key(sheet_name):
    """Clave normalizada para buscar en SHEET_TO_CATEGORY_SLUG."""
    return re.sub(r"\s+", " ", (sheet_name or "").strip()).lower()


def _is_resonator_by_sku(sku):
    """
    True si el part number (SKU) corresponde a un resonador.
    Silenciadores = marcas LTM, GW. Resonadores = LTM con estos prefijos/patrones.
    """
    if not sku or not isinstance(sku, str):
        return False
    u = sku.strip().upper()
    for prefix in RESONATOR_SKU_PREFIXES:
        if u.startswith(prefix.upper()):
            return True
    for sub in RESONATOR_SKU_CONTAINS:
        if sub.upper() in u:
            return True
    return False


def _get_category_for_sheet(sheet_name, euro_norm=None, sku=None):
    """
    Devuelve la categoría (Category) a usar para esta hoja y opcionalmente euro/sku.
    - cataliticos-twc + euro_norm -> subcategoría Euro 3/4/5.
    - Hoja SILENCIADORES: por part number (sku) se distingue Resonadores (LTM según patrón) vs Silenciadores (LTM, GW).
    """
    key = _sheet_key(sheet_name)
    slug = SHEET_TO_CATEGORY_SLUG.get(key)
    if not slug:
        return None
    # Hoja SILENCIADORES: derivar Resonadores vs Silenciadores por part number
    if key == "silenciadores" and sku:
        if _is_resonator_by_sku(sku):
            slug = "resonadores"
        else:
            slug = "silenciadores"
    parent = Category.objects.filter(slug=slug, is_active=True).first()
    if not parent:
        return None
    if slug == "cataliticos-twc" and euro_norm:
        euro_clean = (euro_norm or "").strip().upper().replace(" ", "")
        if euro_clean in ("EURO3", "3"):
            child = Category.objects.filter(slug="cataliticos-twc-euro3", is_active=True).first()
        elif euro_clean in ("EURO4", "4"):
            child = Category.objects.filter(slug="cataliticos-twc-euro4", is_active=True).first()
        elif euro_clean in ("EURO5", "5"):
            child = Category.objects.filter(slug="cataliticos-twc-euro5", is_active=True).first()
        else:
            child = None
        return child if child else parent
    return parent


def _to_decimal(val):
    if val is None:
        return Decimal("0")
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(",", ".")
    try:
        return Decimal(re.sub(r"[^\d.-]", "", s) or "0")
    except Exception:
        return Decimal("0")


def _to_decimal_optional(val):
    """Convierte celda a Decimal; devuelve None si está vacía o no es número."""
    if val is None or (isinstance(val, str) and not str(val).strip()):
        return None
    try:
        return _to_decimal(val)
    except Exception:
        return None


def _find_column_index(header_row, names):
    """Devuelve el índice de la primera columna cuyo valor (normalizado) está en names."""
    header_row = list(header_row) if header_row else []
    names_upper = {str(n).strip().upper() for n in names}
    for i, cell in enumerate(header_row):
        v = (str(cell).strip().upper() if cell else "") or ""
        if v in names_upper:
            return i
    return None


def _extract_dimensions(row, header_row):
    """
    Si la hoja tiene columnas PESO/LARGO/ANCHO/ALTO/VOLUMEN (o equivalentes en inglés),
    extrae sus valores. Devuelve (weight, length, width, height, volume).
    Incluye nombres en inglés para hojas como SILENCIADORES (BODY, OVERALL, etc.).
    """
    row = list(row) if row else []
    w = l = wi = h = vol = None
    if header_row:
        hr = list(header_row)
        idx = _find_column_index(hr, ("PESO", "PESO (KG)", "PESO(KG)", "WEIGHT"))
        if idx is not None and idx < len(row):
            w = _to_decimal_optional(row[idx])
        idx = _find_column_index(hr, ("LARGO", "LARGO (CM)", "LONGITUD", "LENGTH", "OVERALL"))
        if idx is not None and idx < len(row):
            l = _to_decimal_optional(row[idx])
        idx = _find_column_index(hr, ("ANCHO", "ANCHO (CM)", "WIDTH", "BODY"))
        if idx is not None and idx < len(row):
            wi = _to_decimal_optional(row[idx])
        idx = _find_column_index(hr, ("ALTO", "ALTO (CM)", "ALTURA", "HEIGHT"))
        if idx is not None and idx < len(row):
            h = _to_decimal_optional(row[idx])
        idx = _find_column_index(hr, ("VOLUMEN", "VOL", "VOLUME"))
        if idx is not None and idx < len(row):
            vol = _to_decimal_optional(row[idx])
    return w, l, wi, h, vol


# Nombres de columna que no deben importarse como SKU
_HEADER_SKUS = frozenset({"CODIGO", "VALOR", "PART#", "PRICE", "INLET/OUTLET", "CONFIGURE", "BODY", "OVERALL", "MATERIAL", "PHOTO", "MEDIDA", "MODELO", "PULGADAS", "PRECIO", "FLEXIBLES", "COLAS", "REFORZADOS", "EURO", "IVA", "INCLUIDO"})


def _to_sku(val):
    """
    Genera SKU (part number) a partir del valor.
    Para flexibles la medida es el part number: coma decimal → punto (ej. 2,5 X 6 → 2.5X6).
    """
    if val is None:
        return ""
    s = str(val).strip().upper()
    s = s.replace(",", ".")  # coma → punto (medida/part number)
    s = re.sub(r"\s*[xX]\s*", "X", s)  # "2.5 X 6" → "2.5X6"
    s = re.sub(r"\s+", "-", s)[:50]
    if s in _HEADER_SKUS:
        return ""
    return s if s else ""


def _build_product_name(row, code_col, name_cols):
    """Construye nombre del producto desde código y columnas extra."""
    code = (row[code_col] or "").strip() if code_col is not None else ""
    parts = [code] if code else []
    for idx in name_cols:
        if idx is not None and idx < len(row) and row[idx]:
            v = str(row[idx]).strip()
            if v and v not in parts:
                parts.append(v)
    return " ".join(parts)[:255] if parts else (code or "Sin nombre")


def _find_header_row(rows, keywords):
    """Encuentra la fila que contiene alguna de las palabras clave (para código/precio)."""
    for i, row in enumerate(rows):
        row_str = " ".join(str(c).upper() for c in row if c)
        if any(k in row_str for k in keywords):
            return i
    return -1


def _extract_euro(row, header_row):
    """Si el header tiene columna EURO / EURO3 / EURO4 / EURO5, devuelve el valor normalizado."""
    if not header_row:
        return None
    idx = _find_column_index(header_row, ("EURO", "EURO3", "EURO4", "EURO5", "NORMA"))
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    s = str(val).strip().upper().replace(" ", "")
    if "5" in s or s == "EURO5":
        return "EURO5"
    if "4" in s or s == "EURO4":
        return "EURO4"
    if "3" in s or s == "EURO3":
        return "EURO3"
    return None


def _ensure_item_9(item):
    """Asegura que el item tenga 9 elementos (sku, name, price, w, l, wi, h, vol, euro_norm)."""
    while len(item) < 9:
        item = item + (None,)
    return item[:9]


def _parse_sheet(ws):
    """
    Parsea una hoja y devuelve (sheet_title, category_name, list of 9-tuples).
    Cada tuple: (sku, name, price, weight, length, width, height, volume, euro_norm).
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ws.title, _normalize_category_name(ws.title), []

    category_name = _normalize_category_name(ws.title)
    data = []

    ncols = max(len(r) if r else 0 for r in rows) if rows else 0

    if ncols >= 10:
        # SILENCIADORES: PART# col 0, PRICE col 8, nombre desde varias columnas
        header_idx = _find_header_row(rows[:10], ["PART#", "PRICE"])
        if header_idx < 0:
            header_idx = 3
        header_row = rows[header_idx] if header_idx < len(rows) else []
        start = header_idx + 1
        for row in rows[start:]:
            row = list(row) if row else []
            if len(row) < 9:
                continue
            sku = _to_sku(row[0])
            price = _to_decimal(row[8])
            if not sku or price <= 0:
                continue
            name = _build_product_name(row, 0, [1, 2, 3])
            w, ln, wi, h, vol = _extract_dimensions(row, header_row)
            data.append(_ensure_item_9((sku, name, price, w, ln, wi, h, vol, None)))

    elif ncols >= 4:
        # CATALITICOS TWC: codigo, modelo, pulgadas, precio
        header_idx = _find_header_row(rows[:5], ["codigo", "CODIGO", "precio", "PRECIO"])
        if header_idx < 0:
            header_idx = 1
        header_row = rows[header_idx] if header_idx < len(rows) else []
        start = header_idx + 1
        for row in rows[start:]:
            row = list(row) if row else []
            if len(row) < 4:
                continue
            sku = _to_sku(row[0])
            price = _to_decimal(row[3])
            if not sku or price <= 0:
                continue
            name = _build_product_name(row, 0, [1, 2])
            w, ln, wi, h, vol = _extract_dimensions(row, header_row)
            euro_norm = _extract_euro(row, header_row)
            data.append(_ensure_item_9((sku, name, price, w, ln, wi, h, vol, euro_norm)))

    elif ncols >= 3:
        # CATALITICOS CLF o TIPO ORIGINAL: code, ?, price
        header_idx = _find_header_row(rows[:5], ["VALOR", "MEDIDA", "codigo", "CODIGO", "EURO"])
        if header_idx < 0:
            header_idx = 1
        header_row = rows[header_idx] if header_idx < len(rows) else []
        start = header_idx + 1
        for row in rows[start:]:
            row = list(row) if row else []
            if len(row) < 3:
                continue
            sku = _to_sku(row[0])
            price = _to_decimal(row[2])
            if not sku or price <= 0:
                continue
            name = _build_product_name(row, 0, [1])
            w, ln, wi, h, vol = _extract_dimensions(row, header_row)
            euro_norm = _extract_euro(row, header_row)
            data.append(_ensure_item_9((sku, name, price, w, ln, wi, h, vol, euro_norm)))

    else:
        # 2 columnas: FLEXIBLES, COLAS DE ESCAPE
        header_idx = _find_header_row(rows[:6], ["VALOR", "CODIGO", "FLEXIBLES", "COLAS"])
        if header_idx < 0:
            header_idx = 2
        header_row = rows[header_idx] if header_idx < len(rows) else []
        start = header_idx + 1
        for row in rows[start:]:
            row = list(row) if row else []
            if len(row) < 2:
                continue
            code = (row[0] or "").strip()
            price = _to_decimal(row[1])
            if not code or price <= 0:
                continue
            sku = _to_sku(code)
            if not sku:
                code_normalized = code.replace(",", ".")
                sku = slugify(code_normalized).replace("-", "")[:50].upper() or "ITEM"
            name = str(code).replace(",", ".")[:255]
            w, ln, wi, h, vol = _extract_dimensions(row, header_row)
            data.append(_ensure_item_9((sku, name, price, w, ln, wi, h, vol, None)))

    return ws.title, category_name, data


class Command(BaseCommand):
    help = "Carga productos desde lista precios publico.xlsx (por categoría)."

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            type=str,
            help="Ruta al archivo .xlsx (ej: /home/usuario/lista precios publico.xlsx)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Solo mostrar qué se importaría, sin guardar.",
        )

    def handle(self, path, dry_run, **options):
        path = Path(path).resolve()
        if not path.exists():
            # Buscar el archivo en cwd, misc/, raíz del proyecto, y carpeta del comando (management/commands/)
            filename = path.name if path.suffix.lower() == ".xlsx" else "lista precios publico.xlsx"
            commands_dir = Path(__file__).resolve().parent  # apps/catalog/management/commands/
            candidates = [
                Path.cwd() / filename,
                Path.cwd() / "lista precios publico.xlsx",
                Path.cwd() / "misc" / filename,
                Path.cwd() / "misc" / "lista precios publico.xlsx",
                commands_dir / filename,
                commands_dir / "lista precios publico.xlsx",
            ]
            try:
                from django.conf import settings
                base_dir = Path(settings.BASE_DIR)
            except Exception:
                base_dir = Path(__file__).resolve().parents[3]
            candidates.extend([
                base_dir / filename,
                base_dir / "lista precios publico.xlsx",
                base_dir / "misc" / filename,
                base_dir / "misc" / "lista precios publico.xlsx",
            ])
            found = None
            for c in candidates:
                if c.resolve().exists():
                    found = c.resolve()
                    break
            if found:
                path = found
                self.stdout.write(self.style.WARNING(
                    f"No se encontró la ruta indicada. Usando archivo encontrado: {path}"
                ))
            else:
                self.stderr.write(self.style.ERROR(f"Archivo no encontrado: {path}"))
                self.stderr.write("")
                self.stderr.write("Se buscó también en:")
                for c in candidates:
                    self.stderr.write(f"  - {c.resolve()}")
                self.stderr.write("")
                self.stderr.write(
                    "El archivo .xlsx debe estar en este servidor. Opciones:"
                )
                self.stderr.write(
                    "  1. Subir/copiar el Excel al proyecto y usar su ruta absoluta."
                )
                self.stderr.write(
                    "  2. Ver qué .xlsx hay en el directorio actual:"
                )
                self.stderr.write(
                    f'     ls -la "{Path.cwd()}"/*.xlsx'
                )
                self.stderr.write(
                    "  3. Una vez tengas el archivo en el servidor, ejecutar:"
                )
                self.stderr.write(
                    f'     python manage.py load_precios_xlsx "/ruta/real/al/archivo.xlsx"'
                )
                return

        self.stdout.write(f"Leyendo {path} ...")
        wb = openpyxl.load_workbook(path, read_only=True)

        created_cats = 0
        created_prods = 0
        updated_prods = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_title, category_name, items = _parse_sheet(ws)
            if not items:
                self.stdout.write(f"  {sheet_title}: sin filas de datos, omitido.")
                continue

            # Resolver categoría: mapeo a categorías existentes o crear por nombre de hoja
            # Para SILENCIADORES la categoría se resuelve por ítem (part number -> Resonadores vs Silenciadores)
            is_silenciadores_sheet = _sheet_key(sheet_title) == "silenciadores"
            cat = None if is_silenciadores_sheet else _get_category_for_sheet(sheet_title)
            if cat is None and not is_silenciadores_sheet:
                cat, created = Category.objects.get_or_create(
                    name=category_name,
                    defaults={
                        "slug": slugify(category_name),
                        "is_active": True,
                        "parent": None,
                    },
                )
                if created:
                    created_cats += 1
            else:
                created = False

            is_flexibles = _sheet_key(sheet_title) == "flexibles"

            if dry_run:
                cat_label = "Silenciadores/Resonadores (por part number)" if is_silenciadores_sheet else (cat.name if cat else category_name)
                self.stdout.write(f"  [{category_name}] -> {cat_label} ({len(items)} productos)")
                for item in items[:5]:
                    sku, name, price = item[0], item[1], item[2]
                    if is_flexibles:
                        display = get_display_name_for_sku(sku, include_suffix=False)
                        name = (display + FLEXIBLES_SUFFIX)[:255] if display else (name + FLEXIBLES_SUFFIX)[:255]
                    subcat = ""
                    if is_silenciadores_sheet:
                        subcat = " [Resonador]" if _is_resonator_by_sku(sku) else " [Silenciador]"
                    dims = ""
                    if len(item) >= 8 and any(item[i] for i in range(3, 8)):
                        dims = " | " + " ".join(f"{k}={v}" for k, v in zip(("peso", "largo", "ancho", "alto", "vol"), item[3:8]) if v is not None)
                    euro = f" euro={item[8]}" if len(item) > 8 and item[8] else ""
                    self.stdout.write(f"    - {sku} | {name[:40]}... | {price}{subcat}{dims}{euro}")
                if len(items) > 5:
                    self.stdout.write(f"    ... y {len(items) - 5} más")
                continue

            for item in items:
                sku, name, price = item[0], item[1], item[2]
                weight = item[3] if len(item) > 3 else None
                length = item[4] if len(item) > 4 else None
                width = item[5] if len(item) > 5 else None
                height = item[6] if len(item) > 6 else None
                volume = item[7] if len(item) > 7 else None
                euro_norm = item[8] if len(item) > 8 else None
                if not sku:
                    continue

                # Categoría por ítem: TWC por euro; SILENCIADORES por part number (Resonadores vs Silenciadores)
                if is_silenciadores_sheet:
                    product_cat = _get_category_for_sheet(sheet_title, euro_norm=None, sku=sku)
                    if product_cat is None:
                        slug_choice = "resonadores" if _is_resonator_by_sku(sku) else "silenciadores"
                        product_cat, _ = Category.objects.get_or_create(
                            slug=slug_choice,
                            defaults={
                                "name": "Resonadores" if slug_choice == "resonadores" else "Silenciadores de Alto Flujo",
                                "parent": None,
                                "is_active": True,
                            },
                        )
                else:
                    product_cat = _get_category_for_sheet(sheet_title, euro_norm) or cat

                if is_flexibles:
                    display = get_display_name_for_sku(sku, include_suffix=False)
                    if display:
                        name = (display + FLEXIBLES_SUFFIX)[:255]
                    else:
                        name = (name + FLEXIBLES_SUFFIX)[:255]

                slug_base = slugify(sku)[:280]
                slug = slug_base
                cnt = 0
                while Product.objects.filter(slug=slug).exists():
                    cnt += 1
                    slug = f"{slug_base}-{cnt}"[:280]

                defaults = {
                    "name": name or sku,
                    "slug": slug,
                    "category": product_cat,
                    "price": price,
                    "cost_price": Decimal("0"),
                    "stock": 0,
                    "is_active": True,
                    "deleted_at": None,
                }
                if weight is not None:
                    defaults["weight"] = weight
                if length is not None:
                    defaults["length"] = length
                if width is not None:
                    defaults["width"] = width
                if height is not None:
                    defaults["height"] = height
                if volume is not None:
                    defaults["volume"] = volume
                if euro_norm is not None:
                    defaults["euro_norm"] = euro_norm

                prod, created = Product.objects.update_or_create(
                    sku=sku,
                    defaults=defaults,
                )
                if created:
                    created_prods += 1
                else:
                    updated_prods += 1
                prod.refresh_quality(save=True)

        wb.close()

        if dry_run:
            self.stdout.write(self.style.SUCCESS("Dry-run listo. Ejecuta sin --dry-run para importar."))
            return

        self.stdout.write(
            self.style.SUCCESS(
                f"Listo: {created_cats} categorías nuevas, {created_prods} productos nuevos, {updated_prods} actualizados."
            )
        )
